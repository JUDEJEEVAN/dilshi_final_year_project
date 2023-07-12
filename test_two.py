import os
from flask import Flask, render_template, request
import openpyxl

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('test_index_new.html')

@app.route('/merge', methods=['GET', 'POST'])
def merge():
    if request.method == 'POST':
        directory_path = request.form['directory_path']
        file_list = os.listdir(directory_path)
        excel_files = [file for file in file_list if file.endswith(('.xlsx', '.xls'))]
        return render_template('merge.html', excel_files=excel_files)
    return render_template('test_merge_new.html')

@app.route('/merge/process', methods=['POST'])
def merge_process():
    selected_files = request.form.getlist('selected_files')
    selected_file_names = []
    selected_titles_dict = {}
    data_dict_row = {}

    for selected_file in selected_files:
        selected_file_names.append(selected_file)
        wb = openpyxl.load_workbook(os.path.join(directory_path, selected_file))
        sheet = wb.active

        titles = [cell.value for cell in sheet[1]]
        selected_titles = request.form.getlist(selected_file)

        selected_titles = [int(title) for title in selected_titles]
        selected_titles_dict[selected_file] = selected_titles

        row_num = 2
        for row in sheet.iter_rows(min_row=row_num, values_only=True):
            for title_index in selected_titles:
                if row_num not in data_dict_row:
                    data_dict_row[row_num] = [row[title_index]]
                else:
                    data_dict_row[row_num].append(row[title_index])

            row_num += 1

        wb.close()

    header_row = [titles[title_index] for selected_file in selected_file_names
                  for title_index in selected_titles_dict[selected_file]]

    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active
    new_sheet.append(header_row)

    for new_sheet_row, data_row in data_dict_row.items():
        new_sheet.append(data_row)

    new_file_name = request.form['new_file_name'] + '.xlsx'
    new_file_location = request.form['new_file_location']
    new_file_path = os.path.join(new_file_location, new_file_name)
    new_wb.save(new_file_path)
    new_wb.close()

    return render_template('test_merge_process.html', new_file_name=new_file_name)

if __name__ == '__main__':
    app.run()
