import os
import openpyxl

# Prompt the user to enter the directory location of the Excel files
directory_path = input("Enter the directory location of the Excel files: ")
# directory_path = r"F:\PycharmProjects\Stack10-1\InExcel"

# Check if the directory exists
if not os.path.exists(directory_path):
    print("Directory does not exist!")
    exit()

# Get a list of all files in the directory
file_list = os.listdir(directory_path)

# Filter only Excel files
excel_files = []
for file in file_list:
    if file.endswith('.xlsx') or file.endswith('.xls'):
        excel_files.append(file)

# Display the available file names
print("Excel files found within the given directory:")
count = 1
for file in excel_files:
    file_name = os.path.splitext(file)[0]
    print(f"{count}) {file_name}")
    count += 1

# ----------------------------------------------------------------------------------------------

# Prompt the user to select the files to add to the new file
selected_files = input("Select the file(s) that you want to add to the new file (separated by commas): ")
selected_files = selected_files.split(",")

# Store the selected file names
selected_file_names = []
print("\nSelected files:")
for file_index in selected_files:
    file_index = int(file_index.strip()) - 1  # Convert to integer and subtract 1 to get the index
    if file_index < 0 or file_index >= len(excel_files):
        print(f"Invalid file number: {file_index + 1}")
        continue
    selected_file = excel_files[file_index]
    selected_file_names.append(selected_file)
    print(selected_file)

# Create an empty dictionary to store the selected titles for each file
selected_titles_dict = {}

# Iterate over the selected files and display their titles
print("\nViewing the titles within the selected files...")
for selected_file in selected_file_names:
    print(selected_file)
    wb = openpyxl.load_workbook(os.path.join(directory_path, selected_file))
    sheet = wb.active

    titles = []
    for i, cell in enumerate(sheet[1], start=1):
        titles.append(cell.value)
        print(f"{i}) {cell.value}")

    # titles = []
    # i = 1
    # for cell in sheet[1]:
    #     titles.append(cell.value)
    #     print(f'{i}) {cell.value}')
    #     i += 1

    # Prompt the user to select the titles to save to the new Excel file
    selected_titles = input("Select the titles of which you want to save the details (separated by commas): ")
    selected_titles = selected_titles.split(",")
    selected_titles = [int(title.strip()) - 1 for title in selected_titles]  # Convert to integer and subtract 1
    # selected_titles = [int(title.strip()) - 1 for title in selected_titles]

    # Store the selected titles for the current file
    selected_titles_dict[selected_file] = selected_titles

    wb.close()

# Print the overall selected files and titles
print("\nOverall, you have selected the following:")
for selected_file in selected_file_names:
    print(f"\n{selected_file} - ", end="")
    selected_titles = selected_titles_dict[selected_file]
    for title_index in selected_titles:
        print(titles[title_index], end=", ")
    print()

# Prompt the user to confirm merging the data into a new Excel file
merge_data = input("\nDo you want to merge those data and write to a new Excel file? (y/n): ")
if merge_data.lower() != "y":
    print("Data merging cancelled.")
    exit()

# Prompt the user for the name of the new Excel file
new_file_name = input("\nProvide the name for the new Excel file: ")
new_file_name += ".xlsx"

# Prompt the user for the location to store the new file
new_file_location = input("Where do you want to store the newly created file?: ")

# Write the headers to the new file
header_row = []
for selected_file in selected_file_names:
    selected_titles = selected_titles_dict[selected_file]
    for title_index in selected_titles:
        header_row.append(titles[title_index])

# Iterate over the selected files and write their data to the new file
data_dict_row = {}
for selected_file in selected_file_names:
    wb = openpyxl.load_workbook(os.path.join(directory_path, selected_file))
    sheet = wb.active

    # Create dictionary of lists; key is the row number and value is all
    # values from all sheets to be written that row
    row_num = 2
    for row in sheet.iter_rows(min_row=row_num, values_only=True):
        selected_titles = selected_titles_dict[selected_file]
        for title_index in selected_titles:
            if row_num not in data_dict_row:
                data_dict_row[row_num] = [row[title_index]]
            else:
                data_dict_row[row_num] += [row[title_index]]

        row_num += 1
    wb.close()

# Create a new workbook to store the merged data
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet.append(header_row)

# Loop through data_dict_rows for each row list
for new_sheet_row, data_row in data_dict_row.items():
    new_sheet.append(data_row)

# Save the new file
new_file_path = os.path.join(new_file_location, new_file_name)
new_wb.save(new_file_path)
new_wb.close()

print(f"\nData written to \"{new_file_name}\"")